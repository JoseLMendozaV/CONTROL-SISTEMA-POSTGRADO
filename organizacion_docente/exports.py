from io import BytesIO

from django.http import HttpResponse
from django.utils import timezone

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

from .models import EstadoProcesoOrganizacion


COLOR_PRIMARIO = "0F3D5E"
COLOR_SECUNDARIO = "115B83"
COLOR_VERDE = "DCFCE7"
COLOR_AMARILLO = "FEF9C3"
COLOR_ROJO = "FEE2E2"
COLOR_AZUL = "DBEAFE"
COLOR_GRIS = "F1F5F9"
COLOR_BLANCO = "FFFFFF"
COLOR_TEXTO = "1E293B"


def aplicar_estilo_titulo(celda):
    celda.font = Font(bold=True, color=COLOR_BLANCO, size=12)
    celda.fill = PatternFill("solid", fgColor=COLOR_PRIMARIO)
    celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def aplicar_estilo_subtitulo(celda):
    celda.font = Font(bold=True, color=COLOR_TEXTO)
    celda.fill = PatternFill("solid", fgColor=COLOR_GRIS)
    celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def aplicar_bordes(ws):
    borde = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )

    for row in ws.iter_rows():
        for cell in row:
            cell.border = borde
            cell.alignment = Alignment(vertical="center", wrap_text=True)


def ajustar_columnas(ws, ancho_maximo=45):
    for column_cells in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:
            try:
                value = str(cell.value) if cell.value is not None else ""
                if len(value) > max_length:
                    max_length = len(value)
            except Exception:
                pass

        ancho = min(max_length + 3, ancho_maximo)
        ws.column_dimensions[column_letter].width = max(ancho, 12)


def valor_decimal(valor):
    if valor is None:
        return 0
    return float(valor)


def valor_si_no(valor):
    return "Sí" if valor else "No"


def obtener_estado(organizacion, tipo_estado):
    estados = {
        estado.tipo_estado: estado
        for estado in organizacion.estados.all()
    }

    return estados.get(tipo_estado)


def crear_hoja_organizaciones(wb, organizaciones):
    ws = wb.active
    ws.title = "Organizaciones"

    columnas = [
        "No. PAG.",
        "DOCENTE",
        "CEDULA",
        "HORARIO ELABORADO",
        "ORG. DOC ELABORADA",
        "CALENDARIO DE PAGO ELABORADO Y ENVIADO",
        "ORG. DOC ENVIADA A LA VIPE",
        "ORG. DOC. CON No. DE POSICION",
        "ORG. DOC ENVIADA PARA FIRMA ELECTRONICA",
        "FIRMADO POR AUTORIDADES",
        "ORG. DOC. ENVIADA A REC. HUMANOS",
        "ACTA RECIBIDA",
        "ACTA FIRMADA ENVIADA A RH",
        "AÑO",
        "SEMESTRE",
        "FACULTAD",
        "PROGRAMA",
        "GRUPO/AULA",
        "ASIGNATURA",
        "COD. ASIG.",
        "COD. HOR.",
        "TOTAL DE HORAS",
        "FECHAS DE CLASES",
        "HORARIO",
        "FECHA DE MATRICULA",
        "CANT. ESTUD. MATRIC.",
        "TOTAL DE CRÉDITOS",
        "TOTAL LAB.",
        "TOTAL NO EXONERADOS",
        "EXON. 50%",
        "CANT. EXON. 25%",
        "INGRESO POR LAB. (B/.)",
        "TOTAL DE INGRESOS",
        "PAGO DEL DOCENTE (B/.)",
        "UTILIDAD NETA",
        "ESTADO GENERAL",
        "PORCENTAJE DE AVANCE",
        "OBSERVACIONES",
    ]

    ws.append(columnas)

    for cell in ws[1]:
        aplicar_estilo_titulo(cell)

    estados_exportar = [
        "horario_elaborado",
        "organizacion_docente_elaborada",
        "calendario_pago_elaborado_enviado",
        "organizacion_enviada_vipe",
        "organizacion_con_numero_posicion",
        "organizacion_enviada_firma_electronica",
        "firmado_por_autoridades",
        "organizacion_enviada_recursos_humanos",
        "acta_recibida",
        "acta_firmada_enviada_rh",
    ]

    for org in organizaciones:
        estados_valores = []

        for tipo_estado in estados_exportar:
            estado = obtener_estado(org, tipo_estado)
            estados_valores.append(valor_si_no(estado.completado if estado else False))

        fila = [
            org.numero_pago or "",
            org.docente.nombre_completo,
            org.cedula_docente or org.docente.cedula,
            *estados_valores,
            org.anio,
            org.get_semestre_display(),
            str(org.facultad),
            org.programa.nombre,
            org.grupo_aula,
            org.asignatura.nombre,
            org.codigo_asignatura or "",
            org.codigo_horario or "",
            valor_decimal(org.total_horas),
            org.fechas_clases or "",
            org.horario or "",
            org.fecha_matricula,
            org.cantidad_estudiantes_matriculados,
            valor_decimal(org.total_creditos),
            valor_decimal(org.total_laboratorio),
            org.total_no_exonerados,
            org.cantidad_exoneracion_50,
            org.cantidad_exoneracion_25,
            valor_decimal(org.ingreso_laboratorio),
            valor_decimal(org.total_ingresos),
            valor_decimal(org.pago_docente),
            valor_decimal(org.utilidad_neta),
            org.estado_general_texto,
            f"{org.porcentaje_avance}%",
            org.observaciones or "",
        ]

        ws.append(fila)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    aplicar_bordes(ws)
    ajustar_columnas(ws)

    # Formato numérico
    columnas_moneda = ["AF", "AG", "AH", "AI"]
    columnas_decimal = ["V", "AA", "AB"]

    for row in range(2, ws.max_row + 1):
        for col in columnas_moneda:
            ws[f"{col}{row}"].number_format = '"B/." #,##0.00'

        for col in columnas_decimal:
            ws[f"{col}{row}"].number_format = "0.00"

        ws[f"Y{row}"].number_format = "dd/mm/yyyy"

    ws.row_dimensions[1].height = 35

    return ws


def crear_hoja_estados(wb, organizaciones):
    ws = wb.create_sheet("Estados Detallados")

    columnas = [
        "No. Pago",
        "Año",
        "Semestre",
        "Docente",
        "Cédula",
        "Programa",
        "Asignatura",
        "Estado administrativo",
        "Completado",
        "Fecha completado",
        "Usuario que completó",
        "Observación",
    ]

    ws.append(columnas)

    for cell in ws[1]:
        aplicar_estilo_titulo(cell)

    for org in organizaciones:
        for estado in org.estados.all().order_by("id"):
            ws.append([
                org.numero_pago or "",
                org.anio,
                org.get_semestre_display(),
                org.docente.nombre_completo,
                org.cedula_docente or org.docente.cedula,
                org.programa.nombre,
                org.asignatura.nombre,
                estado.get_tipo_estado_display(),
                valor_si_no(estado.completado),
                estado.fecha_completado,
                estado.usuario_completo.username if estado.usuario_completo else "",
                estado.observacion or "",
            ])

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    aplicar_bordes(ws)
    ajustar_columnas(ws)

    for row in range(2, ws.max_row + 1):
        ws[f"J{row}"].number_format = "dd/mm/yyyy hh:mm"

    ws.row_dimensions[1].height = 35

    return ws


def crear_hoja_resumen(wb, organizaciones):
    ws = wb.create_sheet("Resumen")

    total_organizaciones = organizaciones.count()

    total_ingresos = sum([org.total_ingresos or 0 for org in organizaciones])
    total_pagos = sum([org.pago_docente or 0 for org in organizaciones])
    utilidad_neta = sum([org.utilidad_neta or 0 for org in organizaciones])

    completas = 0
    en_proceso = 0
    sin_iniciar = 0
    enviadas = 0

    for org in organizaciones:
        codigo = org.estado_general_codigo

        if codigo == "completo":
            completas += 1
        elif codigo == "en_proceso":
            en_proceso += 1
        elif codigo == "sin_iniciar":
            sin_iniciar += 1
        elif codigo == "enviado":
            enviadas += 1

    datos = [
        ["Resumen de exportación", ""],
        ["Fecha de generación", timezone.localtime(timezone.now()).strftime("%d/%m/%Y %H:%M")],
        ["", ""],
        ["Total de organizaciones", total_organizaciones],
        ["Total de ingresos", valor_decimal(total_ingresos)],
        ["Total de pagos docentes", valor_decimal(total_pagos)],
        ["Utilidad neta", valor_decimal(utilidad_neta)],
        ["", ""],
        ["Proceso completo", completas],
        ["En proceso", en_proceso],
        ["Enviado a otra unidad", enviadas],
        ["Sin iniciar", sin_iniciar],
    ]

    for fila in datos:
        ws.append(fila)

    ws["A1"].font = Font(bold=True, color=COLOR_BLANCO, size=14)
    ws["B1"].font = Font(bold=True, color=COLOR_BLANCO, size=14)
    ws["A1"].fill = PatternFill("solid", fgColor=COLOR_PRIMARIO)
    ws["B1"].fill = PatternFill("solid", fgColor=COLOR_PRIMARIO)

    for row in range(4, 13):
        ws[f"A{row}"].font = Font(bold=True, color=COLOR_TEXTO)
        ws[f"A{row}"].fill = PatternFill("solid", fgColor=COLOR_GRIS)

    for cell in ["B5", "B6", "B7"]:
        ws[cell].number_format = '"B/." #,##0.00'

    aplicar_bordes(ws)
    ajustar_columnas(ws)

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 24

    return ws


def generar_excel_organizaciones(organizaciones):
    """
    Genera archivo Excel con:
    1. Organizaciones
    2. Estados Detallados
    3. Resumen
    """

    wb = Workbook()

    crear_hoja_organizaciones(wb, organizaciones)
    crear_hoja_estados(wb, organizaciones)
    crear_hoja_resumen(wb, organizaciones)

    nombre_archivo = (
        "organizaciones_docentes_"
        f"{timezone.localtime(timezone.now()).strftime('%Y%m%d_%H%M')}.xlsx"
    )

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
    )

    response["Content-Disposition"] = f'attachment; filename="{nombre_archivo}"'

    return response