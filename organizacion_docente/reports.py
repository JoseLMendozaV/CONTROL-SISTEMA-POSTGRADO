from io import BytesIO
from decimal import Decimal

from django.db.models import Count, Sum
from django.http import HttpResponse
from django.template.loader import render_to_string
from django.utils import timezone

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

from .models import CohorteProgramaPostgrado

from django.db.models import Count, Sum, Q
from .models import OrganizacionDocente, ProgramaPostgrado


from decimal import Decimal
from io import BytesIO

from django.db.models import Count, Sum, Q
from django.http import HttpResponse
from django.utils import timezone

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

from .models import OrganizacionDocente


COLOR_PRIMARIO = "0F3D5E"
COLOR_GRIS = "F1F5F9"
COLOR_BLANCO = "FFFFFF"
COLOR_BORDE = "CBD5E1"


def decimal_0(valor):
    return valor or Decimal("0.00")


def contar_pendientes(organizaciones, tipo_estado):
    return organizaciones.filter(
        estados__tipo_estado=tipo_estado,
        estados__completado=False,
    ).distinct().count()


def contar_completados(organizaciones, tipo_estado):
    return organizaciones.filter(
        estados__tipo_estado=tipo_estado,
        estados__completado=True,
    ).distinct().count()


def construir_contexto_reportes(organizaciones):
    """
    Construye todos los datos necesarios para la vista, PDF y Excel.
    """

    totales = organizaciones.aggregate(
        total_organizaciones=Count("id"),
        total_ingresos=Sum("total_ingresos"),
        total_pagos=Sum("pago_docente"),
        utilidad_neta=Sum("utilidad_neta"),
        ingreso_laboratorio=Sum("ingreso_laboratorio"),
    )

    total_organizaciones = totales.get("total_organizaciones") or 0
    total_ingresos = decimal_0(totales.get("total_ingresos"))
    total_pagos = decimal_0(totales.get("total_pagos"))
    utilidad_neta = decimal_0(totales.get("utilidad_neta"))
    ingreso_laboratorio = decimal_0(totales.get("ingreso_laboratorio"))

    pendientes_vipe = contar_pendientes(
        organizaciones,
        "organizacion_enviada_vipe",
    )

    enviadas_vipe = contar_completados(
        organizaciones,
        "organizacion_enviada_vipe",
    )

    pendientes_firma = contar_pendientes(
        organizaciones,
        "organizacion_enviada_firma_electronica",
    )

    firmadas_autoridades = contar_completados(
        organizaciones,
        "firmado_por_autoridades",
    )

    pendientes_rh = contar_pendientes(
        organizaciones,
        "organizacion_enviada_recursos_humanos",
    )

    enviadas_rh = contar_completados(
        organizaciones,
        "organizacion_enviada_recursos_humanos",
    )

    actas_pendientes = contar_pendientes(
        organizaciones,
        "acta_recibida",
    )

    actas_recibidas = contar_completados(
        organizaciones,
        "acta_recibida",
    )

    calendarios_pendientes = contar_pendientes(
        organizaciones,
        "calendario_pago_elaborado_enviado",
    )

    calendarios_enviados = contar_completados(
        organizaciones,
        "calendario_pago_elaborado_enviado",
    )

    por_facultad = list(
        organizaciones
        .values("facultad__nombre", "facultad__siglas")
        .annotate(
            total=Count("id"),
            total_ingresos=Sum("total_ingresos"),
            total_pagos=Sum("pago_docente"),
            utilidad_neta=Sum("utilidad_neta"),
        )
        .order_by("facultad__nombre")
    )

    por_programa = list(
        organizaciones
        .values("programa__nombre", "facultad__siglas")
        .annotate(
            total=Count("id"),
            total_ingresos=Sum("total_ingresos"),
            total_pagos=Sum("pago_docente"),
            utilidad_neta=Sum("utilidad_neta"),
        )
        .order_by("programa__nombre")
    )

    por_docente = list(
        organizaciones
        .values("docente__nombre_completo", "cedula_docente")
        .annotate(
            total=Count("id"),
            total_horas=Sum("total_horas"),
            total_pagos=Sum("pago_docente"),
        )
        .order_by("docente__nombre_completo")
    )

    pagos_por_semestre = list(
        organizaciones
        .values("anio", "semestre")
        .annotate(
            total=Count("id"),
            total_pagos=Sum("pago_docente"),
            total_ingresos=Sum("total_ingresos"),
            utilidad_neta=Sum("utilidad_neta"),
        )
        .order_by("-anio", "semestre")
    )

    pendientes_firma_listado = organizaciones.filter(
        estados__tipo_estado="organizacion_enviada_firma_electronica",
        estados__completado=False,
    ).distinct().order_by("-anio", "semestre", "docente__nombre_completo")

    actas_pendientes_listado = organizaciones.filter(
        estados__tipo_estado="acta_recibida",
        estados__completado=False,
    ).distinct().order_by("-anio", "semestre", "docente__nombre_completo")

    calendarios_pendientes_listado = organizaciones.filter(
        estados__tipo_estado="calendario_pago_elaborado_enviado",
        estados__completado=False,
    ).distinct().order_by("-anio", "semestre", "docente__nombre_completo")

    context = {
        "fecha_generacion": timezone.localtime(timezone.now()),

        "total_organizaciones": total_organizaciones,
        "total_ingresos": total_ingresos,
        "total_pagos": total_pagos,
        "utilidad_neta": utilidad_neta,
        "ingreso_laboratorio": ingreso_laboratorio,

        "pendientes_vipe": pendientes_vipe,
        "enviadas_vipe": enviadas_vipe,
        "pendientes_firma": pendientes_firma,
        "firmadas_autoridades": firmadas_autoridades,
        "pendientes_rh": pendientes_rh,
        "enviadas_rh": enviadas_rh,
        "actas_pendientes": actas_pendientes,
        "actas_recibidas": actas_recibidas,
        "calendarios_pendientes": calendarios_pendientes,
        "calendarios_enviados": calendarios_enviados,

        "por_facultad": por_facultad,
        "por_programa": por_programa,
        "por_docente": por_docente,
        "pagos_por_semestre": pagos_por_semestre,

        "pendientes_firma_listado": pendientes_firma_listado[:50],
        "actas_pendientes_listado": actas_pendientes_listado[:50],
        "calendarios_pendientes_listado": calendarios_pendientes_listado[:50],
    }

    return context


def aplicar_estilo_header(cell):
    cell.font = Font(bold=True, color=COLOR_BLANCO)
    cell.fill = PatternFill("solid", fgColor=COLOR_PRIMARIO)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def aplicar_bordes(ws):
    borde = Border(
        left=Side(style="thin", color=COLOR_BORDE),
        right=Side(style="thin", color=COLOR_BORDE),
        top=Side(style="thin", color=COLOR_BORDE),
        bottom=Side(style="thin", color=COLOR_BORDE),
    )

    for row in ws.iter_rows():
        for cell in row:
            cell.border = borde
            cell.alignment = Alignment(vertical="center", wrap_text=True)


def ajustar_columnas(ws):
    for column_cells in ws.columns:
        max_length = 0
        col = get_column_letter(column_cells[0].column)

        for cell in column_cells:
            value = str(cell.value) if cell.value is not None else ""
            max_length = max(max_length, len(value))

        ws.column_dimensions[col].width = min(max_length + 3, 45)


def escribir_tabla(ws, titulo, headers, filas, fila_inicio):
    ws.merge_cells(
        start_row=fila_inicio,
        start_column=1,
        end_row=fila_inicio,
        end_column=len(headers),
    )

    titulo_cell = ws.cell(row=fila_inicio, column=1)
    titulo_cell.value = titulo
    titulo_cell.font = Font(bold=True, color=COLOR_BLANCO, size=13)
    titulo_cell.fill = PatternFill("solid", fgColor=COLOR_PRIMARIO)

    fila_header = fila_inicio + 1

    for col_index, header in enumerate(headers, start=1):
        cell = ws.cell(row=fila_header, column=col_index)
        cell.value = header
        aplicar_estilo_header(cell)

    fila_actual = fila_header + 1

    for fila in filas:
        for col_index, valor in enumerate(fila, start=1):
            ws.cell(row=fila_actual, column=col_index).value = valor

        fila_actual += 1

    return fila_actual + 2


def generar_excel_reportes(organizaciones):
    """
    Genera Excel con reportes consolidados.
    """

    context = construir_contexto_reportes(organizaciones)

    wb = Workbook()

    # ========================================================
    # Hoja resumen
    # ========================================================
    ws = wb.active
    ws.title = "Resumen"

    resumen = [
        ["Fecha de generación", context["fecha_generacion"].strftime("%d/%m/%Y %H:%M")],
        ["Total de organizaciones", context["total_organizaciones"]],
        ["Total de ingresos", float(context["total_ingresos"])],
        ["Total de pagos docentes", float(context["total_pagos"])],
        ["Utilidad neta", float(context["utilidad_neta"])],
        ["Ingreso laboratorio", float(context["ingreso_laboratorio"])],
        ["Pendientes VIPE", context["pendientes_vipe"]],
        ["Enviadas a VIPE", context["enviadas_vipe"]],
        ["Pendientes firma electrónica", context["pendientes_firma"]],
        ["Firmadas por autoridades", context["firmadas_autoridades"]],
        ["Pendientes RH", context["pendientes_rh"]],
        ["Enviadas a RH", context["enviadas_rh"]],
        ["Actas pendientes", context["actas_pendientes"]],
        ["Actas recibidas", context["actas_recibidas"]],
        ["Calendarios pendientes", context["calendarios_pendientes"]],
        ["Calendarios enviados", context["calendarios_enviados"]],
    ]

    ws.append(["Reporte general de Organización Docente", ""])
    ws["A1"].font = Font(bold=True, color=COLOR_BLANCO, size=14)
    ws["A1"].fill = PatternFill("solid", fgColor=COLOR_PRIMARIO)
    ws["B1"].fill = PatternFill("solid", fgColor=COLOR_PRIMARIO)

    for fila in resumen:
        ws.append(fila)

    for row in range(2, ws.max_row + 1):
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"A{row}"].fill = PatternFill("solid", fgColor=COLOR_GRIS)

    for cell in ["B4", "B5", "B6", "B7"]:
        ws[cell].number_format = '"B/." #,##0.00'

    aplicar_bordes(ws)
    ajustar_columnas(ws)

    # ========================================================
    # Hoja por facultad
    # ========================================================
    ws_fac = wb.create_sheet("Por Facultad")
    ws_fac.append(["Facultad", "Siglas", "Total", "Ingresos", "Pagos", "Utilidad"])

    for cell in ws_fac[1]:
        aplicar_estilo_header(cell)

    for item in context["por_facultad"]:
        ws_fac.append([
            item["facultad__nombre"],
            item["facultad__siglas"] or "",
            item["total"],
            float(decimal_0(item["total_ingresos"])),
            float(decimal_0(item["total_pagos"])),
            float(decimal_0(item["utilidad_neta"])),
        ])

    aplicar_bordes(ws_fac)
    ajustar_columnas(ws_fac)

    for row in range(2, ws_fac.max_row + 1):
        for col in ["D", "E", "F"]:
            ws_fac[f"{col}{row}"].number_format = '"B/." #,##0.00'

    # ========================================================
    # Hoja por programa
    # ========================================================
    ws_prog = wb.create_sheet("Por Programa")
    ws_prog.append(["Facultad", "Programa", "Total", "Ingresos", "Pagos", "Utilidad"])

    for cell in ws_prog[1]:
        aplicar_estilo_header(cell)

    for item in context["por_programa"]:
        ws_prog.append([
            item["facultad__siglas"] or "",
            item["programa__nombre"],
            item["total"],
            float(decimal_0(item["total_ingresos"])),
            float(decimal_0(item["total_pagos"])),
            float(decimal_0(item["utilidad_neta"])),
        ])

    aplicar_bordes(ws_prog)
    ajustar_columnas(ws_prog)

    for row in range(2, ws_prog.max_row + 1):
        for col in ["D", "E", "F"]:
            ws_prog[f"{col}{row}"].number_format = '"B/." #,##0.00'

    # ========================================================
    # Hoja por docente
    # ========================================================
    ws_doc = wb.create_sheet("Por Docente")
    ws_doc.append(["Docente", "Cédula", "Total organizaciones", "Total horas", "Total pagos"])

    for cell in ws_doc[1]:
        aplicar_estilo_header(cell)

    for item in context["por_docente"]:
        ws_doc.append([
            item["docente__nombre_completo"],
            item["cedula_docente"] or "",
            item["total"],
            float(decimal_0(item["total_horas"])),
            float(decimal_0(item["total_pagos"])),
        ])

    aplicar_bordes(ws_doc)
    ajustar_columnas(ws_doc)

    for row in range(2, ws_doc.max_row + 1):
        ws_doc[f"E{row}"].number_format = '"B/." #,##0.00'

    # ========================================================
    # Hoja pagos por semestre
    # ========================================================
    ws_sem = wb.create_sheet("Pagos Semestre")
    ws_sem.append(["Año", "Semestre", "Total", "Ingresos", "Pagos", "Utilidad"])

    for cell in ws_sem[1]:
        aplicar_estilo_header(cell)

    for item in context["pagos_por_semestre"]:
        ws_sem.append([
            item["anio"],
            item["semestre"],
            item["total"],
            float(decimal_0(item["total_ingresos"])),
            float(decimal_0(item["total_pagos"])),
            float(decimal_0(item["utilidad_neta"])),
        ])

    aplicar_bordes(ws_sem)
    ajustar_columnas(ws_sem)

    for row in range(2, ws_sem.max_row + 1):
        for col in ["D", "E", "F"]:
            ws_sem[f"{col}{row}"].number_format = '"B/." #,##0.00'

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"reporte_organizacion_docente_{timezone.now().strftime('%Y%m%d_%H%M')}.xlsx"

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    return response


def generar_pdf_reportes(organizaciones, request=None):
    """
    Genera PDF del reporte consolidado.
    """

    from weasyprint import HTML

    context = construir_contexto_reportes(organizaciones)

    html_string = render_to_string(
        "organizacion_docente/documentos/reporte_general_pdf.html",
        context,
    )

    base_url = request.build_absolute_uri("/") if request else None

    pdf_file = HTML(
        string=html_string,
        base_url=base_url,
    ).write_pdf()

    filename = f"reporte_organizacion_docente_{timezone.now().strftime('%Y%m%d_%H%M')}.pdf"

    response = HttpResponse(
        pdf_file,
        content_type="application/pdf",
    )

    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    return response



from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from django.db.models import Count, Sum, Q
from .models import OrganizacionDocente, ProgramaPostgrado, CohorteProgramaPostgrado


# ============================================================
# Informe de Programas de Postgrado estilo Excel oficial
# ============================================================

def etiqueta_periodo_reporte(tipo_periodo, periodo):
    if tipo_periodo == "SEMESTRE":
        if periodo == "I":
            return "I SEMESTRE"
        if periodo == "II":
            return "II SEMESTRE"
        return "SEMESTRES"

    if tipo_periodo == "CUATRIMESTRE":
        if periodo == "C1":
            return "I CUATRIMESTRE"
        if periodo == "C2":
            return "II CUATRIMESTRE"
        if periodo == "C3":
            return "III CUATRIMESTRE"
        return "CUATRIMESTRES"

    if tipo_periodo == "TRIMESTRE":
        if periodo == "T1":
            return "I TRIMESTRE"
        if periodo == "T2":
            return "II TRIMESTRE"
        if periodo == "T3":
            return "III TRIMESTRE"
        if periodo == "T4":
            return "IV TRIMESTRE"
        return "TRIMESTRES"

    if tipo_periodo == "VERANO":
        return "VERANO"

    return "AÑO COMPLETO"


def periodos_para_ingresos(tipo_periodo):
    if tipo_periodo == "SEMESTRE":
        return [
            ("VERANO", "VERANO"),
            ("I", "I SEMESTRE"),
            ("II", "II SEMESTRE"),
        ]

    if tipo_periodo == "CUATRIMESTRE":
        return [
            ("C1", "I CUATRIMESTRE"),
            ("C2", "II CUATRIMESTRE"),
            ("C3", "III CUATRIMESTRE"),
        ]

    if tipo_periodo == "TRIMESTRE":
        return [
            ("T1", "I TRIMESTRE"),
            ("T2", "II TRIMESTRE"),
            ("T3", "III TRIMESTRE"),
            ("T4", "IV TRIMESTRE"),
        ]

    if tipo_periodo == "VERANO":
        return [
            ("VERANO", "VERANO"),
        ]

    return [
        ("VERANO", "VERANO"),
        ("I", "I SEMESTRE"),
        ("II", "II SEMESTRE"),
        ("C1", "I CUATRIMESTRE"),
        ("C2", "II CUATRIMESTRE"),
        ("C3", "III CUATRIMESTRE"),
        ("T1", "I TRIMESTRE"),
        ("T2", "II TRIMESTRE"),
        ("T3", "III TRIMESTRE"),
        ("T4", "IV TRIMESTRE"),
    ]


def aplicar_estilo_informe_programas(ws):
    thin = Side(style="thin", color="000000")
    borde = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in ws.iter_rows():
        for cell in row:
            cell.border = borde
            cell.alignment = Alignment(
                vertical="center",
                wrap_text=True,
            )

    ws.sheet_view.showGridLines = False


def escribir_titulo_informe(ws, fila, titulo):
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=8)
    cell = ws.cell(row=fila, column=1)
    cell.value = titulo
    cell.font = Font(bold=True, size=12)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    return fila + 1


def escribir_encabezado_centro(ws, fila, titulo):
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=8)
    ws.cell(row=fila, column=1).value = "CENTRO REGIONAL DE CHIRIQUI"
    ws.cell(row=fila, column=1).font = Font(bold=True)
    fila += 1

    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=8)
    ws.cell(row=fila, column=1).value = "UNIDAD DE POSTGRADO"
    ws.cell(row=fila, column=1).font = Font(bold=True)
    fila += 1

    return escribir_titulo_informe(ws, fila, titulo)


def escribir_bloque_organizaciones_informe(
    ws,
    fila,
    titulo,
    organizaciones,
    iniciado=False,
):
    fila = escribir_encabezado_centro(ws, fila, titulo)
    fila += 1

    if iniciado:
        headers = [
            "No.",
            "FACULTAD",
            "PROGRAMA",
            "Matriculados al inicio del programa",
            "Matriculados actualmente",
            "Inicia",
            "Finaliza",
            "OBS.",
        ]
    else:
        headers = [
            "No.",
            "FACULTAD",
            "PROGRAMA",
            "Matriculados Actualmente",
            "Inicia",
            "Finaliza",
            "OBSERVACIÓN",
            "",
        ]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=fila, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    fila += 1

    for index, org in enumerate(organizaciones, start=1):
        facultad = org.facultad.nombre

        programa_nombre = org.programa.nombre

        if org.grupo_aula:
            programa_nombre = f"{programa_nombre} ({org.grupo_aula})"

        ws.cell(row=fila, column=1).value = index
        ws.cell(row=fila, column=2).value = facultad
        ws.cell(row=fila, column=3).value = programa_nombre

        if iniciado:
            ws.cell(row=fila, column=4).value = org.matriculados_inicio_reporte
            ws.cell(row=fila, column=5).value = org.matriculados_actuales_reporte
            ws.cell(row=fila, column=6).value = org.inicio_texto_reporte
            ws.cell(row=fila, column=7).value = org.finaliza_texto_reporte
            ws.cell(row=fila, column=8).value = org.observacion_reporte
        else:
            ws.cell(row=fila, column=4).value = org.matriculados_actuales_reporte
            ws.cell(row=fila, column=5).value = org.inicio_texto_reporte
            ws.cell(row=fila, column=6).value = org.finaliza_texto_reporte
            ws.cell(row=fila, column=7).value = org.observacion_reporte

        fila += 1

    return fila + 3


def obtener_organizaciones_informe_filtradas(
    anio,
    tipo_periodo="TODO",
    periodo="",
    facultad=None,
):
    """
    Obtiene organizaciones docentes marcadas para aparecer
    en el informe anual de programas.
    """

    organizaciones = OrganizacionDocente.objects.filter(
        anio=anio,
        activo=True,
        incluir_en_informe_programas=True,
    ).select_related(
        "facultad",
        "programa",
        "docente",
        "asignatura",
    )

    if facultad:
        organizaciones = organizaciones.filter(facultad=facultad)

    if periodo:
        organizaciones = organizaciones.filter(
            Q(semestre=periodo)
            | Q(periodo_inicio_programa=periodo)
            | Q(periodo_finalizacion_programa=periodo)
        )

    elif tipo_periodo != "TODO":
        codigos_periodo = [
            codigo for codigo, _ in periodos_para_ingresos(tipo_periodo)
        ]

        organizaciones = organizaciones.filter(
            Q(semestre__in=codigos_periodo)
            | Q(periodo_inicio_programa__in=codigos_periodo)
            | Q(periodo_finalizacion_programa__in=codigos_periodo)
        )

    return organizaciones.order_by(
        "facultad__nombre",
        "programa__nombre",
        "grupo_aula",
    )


def ingresos_por_programa_periodo(organizaciones, programa, periodo):
    return organizaciones.filter(
        programa=programa,
        semestre=periodo,
    ).aggregate(
        total=Sum("total_ingresos")
    ).get("total") or Decimal("0.00")


def escribir_hoja_informe_programas(wb, anio, tipo_periodo, periodo, facultad):
    ws = wb.active
    ws.title = f"INFORME DE PROGRAMAS {anio}"

    etiqueta = etiqueta_periodo_reporte(tipo_periodo, periodo)

    organizaciones = obtener_organizaciones_informe_filtradas(
        anio=anio,
        tipo_periodo=tipo_periodo,
        periodo=periodo,
        facultad=facultad,
    )

    continuaron = organizaciones.filter(
        estado_informe_programa="CONTINUA"
    )

    culminados = organizaciones.filter(
        estado_informe_programa="CULMINADO"
    )

    iniciaron = organizaciones.filter(
        estado_informe_programa="INICIADO"
    )

    fila = 1

    fila = escribir_bloque_organizaciones_informe(
        ws,
        fila,
        f"PROGRAMAS QUE CONTINUARON EN EL {anio} - {etiqueta}",
        continuaron,
        iniciado=False,
    )

    fila = escribir_bloque_organizaciones_informe(
        ws,
        fila,
        f"PROGRAMAS CULMINADOS EN EL {anio} - {etiqueta}",
        culminados,
        iniciado=False,
    )

    fila = escribir_bloque_organizaciones_informe(
        ws,
        fila,
        f"PROGRAMAS QUE INICIARON EN EL {anio} - {etiqueta}",
        iniciaron,
        iniciado=True,
    )

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 55
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 25
    ws.column_dimensions["F"].width = 25
    ws.column_dimensions["G"].width = 25
    ws.column_dimensions["H"].width = 18

    for row in range(1, ws.max_row + 1):
        ws.row_dimensions[row].height = 32

    aplicar_estilo_informe_programas(ws)


def escribir_hoja_ingresos_programa(wb, organizaciones, anio, tipo_periodo, periodo, facultad):
    ws = wb.create_sheet("INGRESOS POR PROGRAMA")

    periodos = periodos_para_ingresos(tipo_periodo)

    if periodo:
        periodos = [item for item in periodos if item[0] == periodo]

    total_cols = len(periodos) + 3

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    ws.cell(row=1, column=1).value = "UNIVERSIDAD TECNOLÓGICA DE PANAMÁ"

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
    ws.cell(row=2, column=1).value = "CENTRO REGIONAL DE CHIRIQUÍ"

    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=total_cols)
    ws.cell(row=3, column=1).value = "COORDINACIÓN DE POSTGRADO"

    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=total_cols)
    ws.cell(row=4, column=1).value = f"INFORME DE INGRESOS DE POSTGRADO - AÑO {anio}"

    for row in range(1, 5):
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")

    ws.cell(row=6, column=1).value = "FACULTAD"
    ws.cell(row=6, column=2).value = "PROGRAMA"
    ws.cell(row=6, column=3).value = f"INGRESOS POR PROGRAMAS DE POSTGRADO - AÑO {anio} (B/.)"

    ws.merge_cells(start_row=6, start_column=3, end_row=6, end_column=2 + len(periodos) + 1)

    for col, (_, label) in enumerate(periodos, start=3):
        ws.cell(row=7, column=col).value = label

    total_col = 3 + len(periodos)
    ws.cell(row=7, column=total_col).value = "TOTAL"

    for row in [6, 7]:
        for col in range(1, total_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9EAF7")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    programas = ProgramaPostgrado.objects.filter(
        activo=True
    ).select_related("facultad").order_by(
        "facultad__nombre",
        "nombre",
    )

    if facultad:
        programas = programas.filter(facultad=facultad)

    fila = 8
    total_general_periodos = {codigo: Decimal("0.00") for codigo, _ in periodos}
    total_general = Decimal("0.00")

    facultad_actual = None
    subtotal_periodos = {codigo: Decimal("0.00") for codigo, _ in periodos}
    subtotal_total = Decimal("0.00")

    def escribir_subtotal(nombre_facultad):
        nonlocal fila, subtotal_periodos, subtotal_total

        if not nombre_facultad:
            return

        ws.cell(row=fila, column=2).value = f"SUBTOTAL - {nombre_facultad}"
        ws.cell(row=fila, column=2).font = Font(bold=True)

        for idx, (codigo, _) in enumerate(periodos, start=3):
            ws.cell(row=fila, column=idx).value = float(subtotal_periodos[codigo])
            ws.cell(row=fila, column=idx).font = Font(bold=True)

        ws.cell(row=fila, column=total_col).value = float(subtotal_total)
        ws.cell(row=fila, column=total_col).font = Font(bold=True)

        fila += 1

        subtotal_periodos = {codigo: Decimal("0.00") for codigo, _ in periodos}
        subtotal_total = Decimal("0.00")

    for programa in programas:
        nombre_facultad = programa.facultad.siglas or programa.facultad.nombre

        if facultad_actual and facultad_actual != nombre_facultad:
            escribir_subtotal(facultad_actual)

        if facultad_actual != nombre_facultad:
            ws.cell(row=fila, column=1).value = nombre_facultad
            facultad_actual = nombre_facultad

        ws.cell(row=fila, column=2).value = programa.nombre

        total_programa = Decimal("0.00")

        for idx, (codigo_periodo, _) in enumerate(periodos, start=3):
            monto = ingresos_por_programa_periodo(
                organizaciones,
                programa,
                codigo_periodo,
            )

            ws.cell(row=fila, column=idx).value = float(monto)

            subtotal_periodos[codigo_periodo] += monto
            total_general_periodos[codigo_periodo] += monto
            total_programa += monto

        ws.cell(row=fila, column=total_col).value = float(total_programa)

        subtotal_total += total_programa
        total_general += total_programa

        fila += 1

    escribir_subtotal(facultad_actual)

    ws.cell(row=fila, column=2).value = "TOTAL DE INGRESOS*"
    ws.cell(row=fila, column=2).font = Font(bold=True)

    for idx, (codigo, _) in enumerate(periodos, start=3):
        ws.cell(row=fila, column=idx).value = float(total_general_periodos[codigo])
        ws.cell(row=fila, column=idx).font = Font(bold=True)

    ws.cell(row=fila, column=total_col).value = float(total_general)
    ws.cell(row=fila, column=total_col).font = Font(bold=True)

    fila += 1

    ws.cell(row=fila, column=1).value = (
        "* Las cifras de ingresos mostradas son netas, luego de restar los gastos "
        "de honorarios correspondientes"
    )
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=total_col)

    for row in range(8, fila):
        for col in range(3, total_col + 1):
            ws.cell(row=row, column=col).number_format = '"B/." #,##0.00'

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 60

    for col in range(3, total_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16

    for row in range(1, fila + 1):
        ws.row_dimensions[row].height = 25

    aplicar_estilo_informe_programas(ws)


def generar_excel_informe_programas(anio, tipo_periodo="TODO", periodo="", facultad=None):
    """
    Genera Excel con dos hojas:
    1. Informe de programas
    2. Ingresos por programa
    """

    wb = Workbook()

    organizaciones = OrganizacionDocente.objects.filter(
        anio=anio,
        activo=True,
    ).select_related(
        "facultad",
        "programa",
        "docente",
        "asignatura",
    )

    if facultad:
        organizaciones = organizaciones.filter(facultad=facultad)

    if periodo:
        organizaciones = organizaciones.filter(semestre=periodo)

    escribir_hoja_informe_programas(
        wb=wb,
        anio=anio,
        tipo_periodo=tipo_periodo,
        periodo=periodo,
        facultad=facultad,
    )

    escribir_hoja_ingresos_programa(
        wb=wb,
        organizaciones=organizaciones,
        anio=anio,
        tipo_periodo=tipo_periodo,
        periodo=periodo,
        facultad=facultad,
    )

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"informe_programas_postgrado_{anio}_{tipo_periodo.lower()}.xlsx"

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    return response



# ============================================================
# Reporte de posiciones a utilizar
# ============================================================

def etiqueta_periodo_posicion(codigo, anio):
    etiquetas = {
        "VERANO": f"Verano-{anio}",
        "I": f"I Sem -{anio}",
        "II": f"II Sem-{anio}",
        "T1": f"I Trim-{anio}",
        "T2": f"II Trim-{anio}",
        "T3": f"III Trim-{anio}",
        "T4": f"IV Trim-{anio}",
        "C1": f"I Cuat-{anio}",
        "C2": f"II Cuat-{anio}",
        "C3": f"III Cuat-{anio}",
        "ESPECIAL": f"Especial-{anio}",
    }

    return etiquetas.get(codigo, f"{codigo}-{anio}")


def etiqueta_total_periodo_posicion(codigo, anio):
    etiquetas = {
        "VERANO": f"Total Verano {anio}",
        "I": f"Total I Sem-{anio}",
        "II": f"Total II Sem-{anio}",
        "T1": f"Total I Trim-{anio}",
        "T2": f"Total II Trim-{anio}",
        "T3": f"Total III Trim-{anio}",
        "T4": f"Total IV Trim-{anio}",
        "C1": f"Total I Cuat-{anio}",
        "C2": f"Total II Cuat-{anio}",
        "C3": f"Total III Cuat-{anio}",
        "ESPECIAL": f"Total Especial-{anio}",
    }

    return etiquetas.get(codigo, f"Total {codigo}-{anio}")


def orden_periodos_posiciones():
    return [
        "VERANO",
        "I",
        "II",
        "T1",
        "T2",
        "T3",
        "T4",
        "C1",
        "C2",
        "C3",
        "ESPECIAL",
    ]


def monto_posicion_por_tipo(tipo_posicion):
    valores = {
        "16": Decimal("640.00"),
        "32": Decimal("1280.00"),
        "48": Decimal("1920.00"),
        "64": Decimal("2560.00"),
    }

    return valores.get(str(tipo_posicion), Decimal("0.00"))


def obtener_organizaciones_posiciones(anio, periodo="", facultad=None):
    organizaciones = OrganizacionDocente.objects.filter(
        anio=anio,
        activo=True,
        incluir_en_reporte_posiciones=True,
    ).select_related(
        "facultad",
        "programa",
        "docente",
        "asignatura",
    )

    if periodo:
        organizaciones = organizaciones.filter(semestre=periodo)

    if facultad:
        organizaciones = organizaciones.filter(facultad=facultad)

    return organizaciones.order_by(
        "semestre",
        "facultad__nombre",
        "programa__nombre",
    )


def construir_resumen_posiciones(organizaciones):
    resumen = {}

    for org in organizaciones:
        periodo = org.semestre
        facultad = org.facultad_reporte_posicion
        tipo = str(org.tipo_posicion or "")

        if not tipo:
            continue

        key = (periodo, facultad)

        if key not in resumen:
            resumen[key] = {
                "16": 0,
                "32": 0,
                "48": 0,
                "64": 0,
            }

        resumen[key][tipo] += org.cantidad_posiciones_a_utilizar or 0

    return resumen


def escribir_celda_posiciones(ws, fila, col, valor, bold=False, fill=None, align="center"):
    cell = ws.cell(row=fila, column=col)
    cell.value = valor
    cell.alignment = Alignment(
        horizontal=align,
        vertical="center",
        wrap_text=True,
    )

    if bold:
        cell.font = Font(bold=True)

    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)

    return cell


def aplicar_estilo_posiciones(ws):
    thin = Side(style="thin", color="000000")
    borde = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in ws.iter_rows():
        for cell in row:
            cell.border = borde
            cell.alignment = Alignment(
                vertical="center",
                wrap_text=True,
            )

    ws.sheet_view.showGridLines = False


def generar_excel_reporte_posiciones(anio, periodo="", facultad=None):
    organizaciones = obtener_organizaciones_posiciones(
        anio=anio,
        periodo=periodo,
        facultad=facultad,
    )

    resumen = construir_resumen_posiciones(organizaciones)

    wb = Workbook()
    ws = wb.active
    ws.title = f"POSICIONES UTILIZADAS {anio}"

    # Títulos
    ws.merge_cells("A1:K1")
    ws.merge_cells("A2:K2")
    ws.merge_cells("A3:K3")

    escribir_celda_posiciones(
        ws,
        1,
        1,
        "UNIVERSIDAD TECNOLÓGICA DE PANAMÁ",
        bold=True,
    )
    escribir_celda_posiciones(
        ws,
        2,
        1,
        "CENTRO REGIONAL DE CHIRIQUÍ",
        bold=True,
    )
    escribir_celda_posiciones(
        ws,
        3,
        1,
        f"POSICIONES A UTILIZAR EN EL {anio}",
        bold=True,
    )

    # Encabezado principal
    headers_row_4 = [
        "PERIODO",
        "FACULTAD",
        "POSICIONES (B/.)",
        "Total B/.",
        "",
        "Total B/.",
        "",
        "Total B/.",
        "",
        "Total B/.",
        "TOTAL B/.",
    ]

    headers_row_5 = [
        "",
        "",
        640,
        "",
        1280,
        "",
        1920,
        "",
        2560,
        "",
        "",
    ]

    headers_row_6 = [
        "",
        "",
        "(16 horas)",
        "",
        "(32 horas)",
        "",
        "(48 horas)",
        "",
        "(64 horas)",
        "",
        "",
    ]

    for col, value in enumerate(headers_row_4, start=1):
        escribir_celda_posiciones(ws, 4, col, value, bold=True, fill="D9EAD3")

    for col, value in enumerate(headers_row_5, start=1):
        escribir_celda_posiciones(ws, 5, col, value, bold=True, fill="D9EAD3")

    for col, value in enumerate(headers_row_6, start=1):
        escribir_celda_posiciones(ws, 6, col, value, bold=True, fill="D9EAD3")

    fila = 7

    total_general = {
        "16": 0,
        "32": 0,
        "48": 0,
        "64": 0,
    }

    periodos = orden_periodos_posiciones()

    if periodo:
        periodos = [periodo]

    for codigo_periodo in periodos:
        filas_periodo = [
            (per, fac, datos)
            for (per, fac), datos in resumen.items()
            if per == codigo_periodo
        ]

        if not filas_periodo:
            continue

        subtotal_periodo = {
            "16": 0,
            "32": 0,
            "48": 0,
            "64": 0,
        }

        for per, fac, datos in sorted(filas_periodo, key=lambda x: x[1]):
            etiqueta_periodo = etiqueta_periodo_posicion(per, anio)

            total_16 = datos["16"] * 640
            total_32 = datos["32"] * 1280
            total_48 = datos["48"] * 1920
            total_64 = datos["64"] * 2560
            total_fila = total_16 + total_32 + total_48 + total_64

            valores = [
                etiqueta_periodo,
                fac,
                datos["16"],
                total_16,
                datos["32"],
                total_32,
                datos["48"],
                total_48,
                datos["64"],
                total_64,
                total_fila,
            ]

            for col, value in enumerate(valores, start=1):
                align = "left" if col in [1, 2] else "center"
                escribir_celda_posiciones(ws, fila, col, value, align=align)

            for tipo in ["16", "32", "48", "64"]:
                subtotal_periodo[tipo] += datos[tipo]
                total_general[tipo] += datos[tipo]

            fila += 1

        total_16 = subtotal_periodo["16"] * 640
        total_32 = subtotal_periodo["32"] * 1280
        total_48 = subtotal_periodo["48"] * 1920
        total_64 = subtotal_periodo["64"] * 2560
        total_periodo = total_16 + total_32 + total_48 + total_64

        valores_total = [
            etiqueta_total_periodo_posicion(codigo_periodo, anio),
            f"TOTAL - {etiqueta_periodo_posicion(codigo_periodo, anio).upper()}",
            subtotal_periodo["16"],
            total_16,
            subtotal_periodo["32"],
            total_32,
            subtotal_periodo["48"],
            total_48,
            subtotal_periodo["64"],
            total_64,
            total_periodo,
        ]

        for col, value in enumerate(valores_total, start=1):
            align = "left" if col in [1, 2] else "center"
            escribir_celda_posiciones(
                ws,
                fila,
                col,
                value,
                bold=True,
                fill="FFF2CC",
                align=align,
            )

        fila += 1

    # Total anual
    total_16 = total_general["16"] * 640
    total_32 = total_general["32"] * 1280
    total_48 = total_general["48"] * 1920
    total_64 = total_general["64"] * 2560
    total_anual = total_16 + total_32 + total_48 + total_64

    valores_anual = [
        f"TOTAL - {anio}",
        "",
        total_general["16"],
        total_16,
        total_general["32"],
        total_32,
        total_general["48"],
        total_48,
        total_general["64"],
        total_64,
        total_anual,
    ]

    for col, value in enumerate(valores_anual, start=1):
        align = "left" if col in [1, 2] else "center"
        escribir_celda_posiciones(
            ws,
            fila,
            col,
            value,
            bold=True,
            fill="D9EAD3",
            align=align,
        )

    fila += 2

    # Resumen inferior
    posiciones_solicitadas = sum(
        org.cantidad_posiciones_asignadas or 0
        for org in organizaciones
        if org.origen_posicion == "SOLICITADA"
    )

    monto_solicitadas = sum(
        Decimal(org.cantidad_posiciones_asignadas or 0) * Decimal(org.monto_base_posicion or 0)
        for org in organizaciones
        if org.origen_posicion == "SOLICITADA"
    )

    posiciones_reasignadas = sum(
        org.cantidad_posiciones_asignadas or 0
        for org in organizaciones
        if org.origen_posicion == "REASIGNADA"
    )

    monto_reasignadas = sum(
        Decimal(org.cantidad_posiciones_asignadas or 0) * Decimal(org.monto_base_posicion or 0)
        for org in organizaciones
        if org.origen_posicion == "REASIGNADA"
    )

    total_asignadas = posiciones_solicitadas + posiciones_reasignadas
    monto_total_asignadas = monto_solicitadas + monto_reasignadas

    posiciones_utilizar = sum(
        org.cantidad_posiciones_a_utilizar or 0
        for org in organizaciones
    )

    monto_utilizar = sum(
        Decimal(org.cantidad_posiciones_a_utilizar or 0) * Decimal(org.monto_base_posicion or 0)
        for org in organizaciones
    )

    posiciones_sin_utilizar = total_asignadas - posiciones_utilizar
    monto_sin_utilizar = monto_total_asignadas - monto_utilizar

    escribir_celda_posiciones(ws, fila, 2, "Cant. Posiciones", bold=True)
    escribir_celda_posiciones(ws, fila, 3, "Cantidad B/.", bold=True)

    escribir_celda_posiciones(
        ws,
        fila,
        6,
        "* Las posiciones a utilizar incluyen:",
        bold=True,
        align="left",
    )

    fila += 1

    resumen_final = [
        ["Posiciones solicitadas", posiciones_solicitadas, float(monto_solicitadas)],
        ["Posiciones reasignadas", posiciones_reasignadas, float(monto_reasignadas)],
        ["Total de posiciones asignadas", total_asignadas, float(monto_total_asignadas)],
        [f"Posiciones a Utilizar - Año {anio} *", posiciones_utilizar, float(monto_utilizar)],
        ["Posiciones sin utilizar", posiciones_sin_utilizar, float(monto_sin_utilizar)],
    ]

    detalle_tipos = [
        f"{total_general['16']} posiciones de 16 horas (B/. 640.00)",
        f"{total_general['32']} posiciones de 32 horas (B/. 1,280.00)",
        f"{total_general['48']} posiciones de 48 horas (B/. 1,920.00)",
        f"{total_general['64']} posiciones de 64 horas (B/. 2,560.00)",
        "** Posiciones de programas nuevos",
    ]

    for index, row in enumerate(resumen_final):
        escribir_celda_posiciones(ws, fila, 1, row[0], align="left")
        escribir_celda_posiciones(ws, fila, 2, row[1])
        escribir_celda_posiciones(ws, fila, 3, row[2])
        escribir_celda_posiciones(ws, fila, 6, detalle_tipos[index], align="left")
        fila += 1

    # Formatos
    for row in range(7, fila + 1):
        for col in [4, 6, 8, 10, 11, 3]:
            ws.cell(row=row, column=col).number_format = '"B/." #,##0.00'

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 16
    ws.column_dimensions["G"].width = 16
    ws.column_dimensions["H"].width = 16
    ws.column_dimensions["I"].width = 16
    ws.column_dimensions["J"].width = 16
    ws.column_dimensions["K"].width = 16

    for row in range(1, fila + 1):
        ws.row_dimensions[row].height = 24

    aplicar_estilo_posiciones(ws)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"reporte_posiciones_a_utilizar_{anio}.xlsx"

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    return response